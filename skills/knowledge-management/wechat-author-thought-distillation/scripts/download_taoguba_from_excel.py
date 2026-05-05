#!/usr/bin/env python3
"""
淘股吧文章批量下载脚本 - 从Excel提取并下载最近三年文章

使用方法：
    python3 download_taoguba_from_excel.py

前置条件：
    - Excel文件：/Users/me/Downloads/微信公众号文章 (1).xlsx
    - 依赖：pip3 install openpyxl pandas

输出：
    - 链接列表：~/.hermes/taoguba_links_2024_2026.txt
    - 文章目录：~/.hermes/taoguba_corpus_2024_2026/
"""

import json
import time
import urllib.request
import urllib.error
from pathlib import Path
from datetime import datetime, timedelta
import re

# 配置
EXCEL_FILE = Path("/Users/me/Downloads/微信公众号文章 (1).xlsx")
OUTPUT_DIR = Path.home() / ".hermes" / "taoguba_corpus_2024_2026"
LINKS_FILE = Path.home() / ".hermes" / "taoguba_links_2024_2026.txt"
PROGRESS_FILE = OUTPUT_DIR / "download_progress.json"
LOG_FILE = OUTPUT_DIR / "download.log"

# 创建输出目录
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def log(message):
    """记录日志"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_line = f"[{timestamp}] {message}"
    print(log_line)
    
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(log_line + '\n')

def extract_recent_articles():
    """从Excel提取最近三年文章链接"""
    try:
        import openpyxl
        
        log("="*80)
        log("📊 从Excel提取最近三年文章")
        log("="*80)
        
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        ws = wb.active
        
        # 读取表头
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        log(f"✅ 表头: {headers[:10]}")
        
        # 找到关键列
        link_col = date_col = title_col = None
        for i, h in enumerate(headers):
            if h and '链接' in str(h):
                link_col = i
            if h and '发布时间' in str(h):
                date_col = i
            if h and '标题' in str(h):
                title_col = i
        
        log(f"✅ 找到列 - 链接:{link_col}, 日期:{date_col}, 标题:{title_col}")
        
        # 计算三年前的日期
        three_years_ago = datetime.now() - timedelta(days=3*365)
        log(f"📅 筛选范围: {three_years_ago.strftime('%Y-%m-%d')} 至今")
        
        # 读取数据
        articles = []
        total_rows = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_rows += 1
            
            if link_col is not None and row[link_col]:
                # 解析日期
                date_str = str(row[date_col]) if date_col is not None else ''
                
                # 筛选2024-2026年
                if '2024' in date_str or '2025' in date_str or '2026' in date_str:
                    articles.append({
                        'link': str(row[link_col]),
                        'date': date_str[:10],
                        'title': str(row[title_col]) if title_col is not None else ''
                    })
        
        wb.close()
        
        log(f"\n✅ 提取到 {len(articles)} 篇文章（最近三年）")
        log(f"   总数据量: {total_rows} 篇")
        
        # 保存链接列表
        with open(LINKS_FILE, 'w', encoding='utf-8') as f:
            for article in articles:
                f.write(f"{article['link']}\n")
        
        log(f"✅ 已保存链接列表: {LINKS_FILE}")
        
        return articles
        
    except Exception as e:
        log(f"❌ 提取失败: {e}")
        return None

def extract_article_content(url):
    """提取文章内容"""
    try:
        req = urllib.request.Request(url)
        req.add_header('User-Agent', 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36')
        
        with urllib.request.urlopen(req, timeout=30) as response:
            html = response.read().decode('utf-8')
            
            # 提取标题
            title_match = re.search(r'<meta property="og:title" content="([^"]+)"', html)
            title = title_match.group(1) if title_match else "未命名"
            
            # 提取发布时间
            time_match = re.search(r'<meta property="article:published_time" content="([^"]+)"', html)
            publish_time = time_match.group(1) if time_match else datetime.now().isoformat()
            
            # 提取正文（简化版）
            content_match = re.search(r'<div class="rich_media_content[^"]*"[^>]*>(.*?)</div>', html, re.DOTALL)
            if content_match:
                content_html = content_match.group(1)
                content = re.sub(r'<[^>]+>', '', content_html).strip()
            else:
                content = "无法提取内容"
            
            return {
                "title": title,
                "publish_time": publish_time,
                "url": url,
                "content": content[:15000]
            }
    except Exception as e:
        log(f"  ❌ 提取失败: {e}")
        return None

def save_article(article, index):
    """保存文章为Markdown"""
    if not article:
        return False
    
    safe_title = re.sub(r'[<>:"/\\|?*]', '', article['title'])
    filename = f"{index:04d}_{safe_title[:50]}.md"
    filepath = OUTPUT_DIR / filename
    
    md_content = f"""# {article['title']}

**发布时间**: {article['publish_time']}
**原文链接**: {article['url']}

---

{article['content']}

---
*下载时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*
"""
    
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(md_content)
        return True
    except Exception as e:
        log(f"  ❌ 保存失败: {e}")
        return False

def load_progress():
    """加载下载进度"""
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"downloaded": [], "failed": [], "last_index": 0}

def save_progress(progress):
    """保存下载进度"""
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)

def main():
    """主函数"""
    # 步骤1：提取最近三年文章
    articles = extract_recent_articles()
    
    if not articles:
        log("❌ 无法提取文章，退出")
        return
    
    # 步骤2：加载链接列表
    with open(LINKS_FILE, 'r') as f:
        links = [line.strip() for line in f if line.strip()]
    
    total = len(links)
    
    log("\n" + "="*80)
    log("🚀 开始批量下载")
    log("="*80)
    log(f"📊 总链接数: {total}")
    
    # 步骤3：加载进度
    progress = load_progress()
    start_index = progress.get('last_index', 0)
    
    if start_index > 0:
        log(f"🔄 断点续传: 从第 {start_index + 1} 篇开始")
    
    # 步骤4：开始下载
    success_count = len(progress['downloaded'])
    failed_count = len(progress['failed'])
    
    for i in range(start_index, total):
        url = links[i]
        
        log(f"\n[{i+1}/{total}] 下载: {url[:60]}...")
        
        article = extract_article_content(url)
        
        if article:
            if save_article(article, i+1):
                success_count += 1
                progress['downloaded'].append(url)
                log(f"  ✅ 成功 ({success_count}/{total})")
            else:
                failed_count += 1
                progress['failed'].append(url)
        else:
            failed_count += 1
            progress['failed'].append(url)
        
        # 更新进度
        progress['last_index'] = i + 1
        save_progress(progress)
        
        # 控制频率
        time.sleep(1.5)
        
        # 每100篇输出统计
        if (i + 1) % 100 == 0:
            log(f"\n📊 阶段统计: 成功 {success_count}, 失败 {failed_count}, 进度 {(i+1)/total*100:.1f}%")
    
    # 完成
    log("\n" + "="*80)
    log("✅ 下载完成！")
    log(f"  成功: {success_count}")
    log(f"  失败: {failed_count}")
    log(f"  成功率: {success_count/total*100:.1f}%")
    log("="*80)

if __name__ == "__main__":
    main()
